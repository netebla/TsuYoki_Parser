#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Парсер для сайта tsuyoki.ru.
Обрабатывает список артикулов (из XML или Excel). Изображения: при `--local-lures` — сначала
локальная папка моделей, иначе только GitHub (ветка `main`, каталог `TsuYoki Lures 2014-2026`),
при отсутствии — tsuyoki.ru.
Выравнивание: квадрат, воблер по центру и горизонтально (align_image + fallback PIL).
"""

from __future__ import annotations

import logging

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

import math
import os
import random
import re
import shutil
import sys
import tempfile
import time
import xml.etree.ElementTree as ET
from pathlib import Path, PurePosixPath
from typing import Optional
from urllib.parse import quote, urljoin, urlparse

logger.info("Подключаю requests / urllib3 (на медленном диске может занять 1–3+ мин)…")
import requests
from bs4 import BeautifulSoup

logger.info("Подключаю OpenCV и NumPy (холодный старт часто 1–3 мин)…")
try:
    import cv2
    import numpy as np
    _ALIGN_AVAILABLE = True
except (ImportError, OSError) as e:
    cv2 = None
    np = None
    _ALIGN_AVAILABLE = False
    msg = str(e)
    if "mmap" in msg or "dlopen" in msg or "cv2.abi3" in msg:
        logger.warning(
            "OpenCV не загрузился (mmap/dlopen). Частая причина — venv или проект в iCloud Drive "
            "с «оптимизацией хранилища»: бинарник .so не полностью локальный. "
            "Сделайте: перенос проекта на локальный каталог (например ~/Projects/…), "
            "или «Загрузить сейчас» для папки venv, затем "
            "`pip install --force-reinstall --no-cache-dir opencv-python-headless numpy`. "
            "Ошибка: %s",
            msg,
        )
    else:
        logger.warning("OpenCV/NumPy не загрузились, выравнивание отключено: %s", msg)

logger.info("Подключаю Pillow…")
try:
    from PIL import Image
    _PIL_AVAILABLE = True
except ImportError:
    Image = None
    _PIL_AVAILABLE = False

logger.info("Импорт завершён.")


def configure_runtime_logging(*, log_file: Optional[Path] = None, verbose: bool = False) -> None:
    """Доп. настройка логов: файл и/или DEBUG (вызывать из main() после parse_args)."""
    root = logging.getLogger()
    root.setLevel(logging.DEBUG if verbose else logging.INFO)
    use_ts = bool(log_file) or verbose
    fmt = (
        logging.Formatter("%(asctime)s %(levelname)s %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
        if use_ts
        else logging.Formatter("%(message)s")
    )
    for h in root.handlers:
        h.setFormatter(fmt)
        h.setLevel(logging.DEBUG if verbose else logging.INFO)
    if log_file:
        p = Path(log_file).expanduser()
        if not p.is_absolute():
            p = Path.cwd() / p
        p = p.resolve()
        p.parent.mkdir(parents=True, exist_ok=True)
        fh = logging.FileHandler(p, encoding="utf-8")
        fh.setFormatter(fmt)
        fh.setLevel(logging.DEBUG if verbose else logging.INFO)
        root.addHandler(fh)
        logger.info("Лог также пишется в файл: %s", p)
    if verbose:
        logging.getLogger("urllib3").setLevel(logging.WARNING)
        logging.getLogger("requests").setLevel(logging.WARNING)


# ---------------------------------------------------------------------------
# Поиск изображений (по модели и коду цвета из наименования)
# ---------------------------------------------------------------------------

DEFAULT_LURES_ROOT = Path(__file__).resolve().parent / "TsuYoki Lures 2014-2026"
DEFAULT_GITHUB_REPO = "netebla/TsuYoki_Parser"
DEFAULT_GITHUB_BRANCH = "main"
DEFAULT_GITHUB_LURES_DIR = "TsuYoki Lures 2014-2026"


def _normalize_model_for_folder(folder_name: str) -> str:
    """Приводит имя папки к базовой модели для сопоставления."""
    s = folder_name.strip()
    s = re.sub(r"\s*\(\d+\)\s*$", "", s)
    s = re.sub(r"\s*-\s*дополнение\s*$", "", s, flags=re.I)
    s = re.sub(r"\s+2025\s*$", "", s)
    s = re.sub(r"\s*\(НОВАЯ МОДЕЛЬ\)\s*$", "", s, flags=re.I)
    return s.strip().upper()


# Папки вроде «MACHO SR42F» vs наименование «MACHO SR 42F …»
_SR_MR_DIGIT_RE = re.compile(r"\b(SR|MR)\s+(\d)", re.I)


def _lure_model_index_key(name: str) -> str:
    """Ключ индекса: SR/MR + пробел + цифра → без пробела (SR42F и SR 42F в одном бакете)."""
    norm = _normalize_model_for_folder(name)
    if not norm:
        return ""
    return _SR_MR_DIGIT_RE.sub(r"\1\2", norm).upper()


def _lure_model_lookup_keys(name_or_model: str) -> list[str]:
    """Ключи для поиска: канонический, как в названии, и полное схлопывание пробелов."""
    norm = _normalize_model_for_folder(name_or_model)
    if not norm:
        return []
    primary = _SR_MR_DIGIT_RE.sub(r"\1\2", norm).upper()
    keys: list[str] = []
    for k in (primary, norm, re.sub(r"\s+", "", norm)):
        if k and k not in keys:
            keys.append(k)
    return keys


def _extract_color_code_from_filename(filename: str) -> Optional[str]:
    """Извлекает код цвета из имени файла (последний токен)."""
    base = Path(filename).stem
    if not base:
        return None
    parts = base.split()
    if not parts:
        return None
    last = parts[-1]
    if re.match(r"^[\dA-Za-z\-]+$", last) and last not in ("TSUYOKI", "ВОБЛЕР"):
        return last
    return None


_MODEL_SUFFIXES = frozenset({"SP", "SS", "F", "S", "SF", "MR", "SR", "SOFT", "R"})


def _looks_like_color_code(token: str) -> bool:
    """Токен похож на код цвета (041, 013S, FRU), а не на часть модели."""
    if not token:
        return False
    t = token.upper()
    if t in _MODEL_SUFFIXES:
        return False
    if re.match(r"^\d{2,4}[A-Za-z]{0,2}$", token):
        return True
    if re.match(r"^[A-Za-z]+\d*$", token) and len(token) <= 5:
        return True
    if re.match(r"^[\w]+-\d+$", token):
        return True
    return False


def _model_from_name_tokens(tokens: list) -> Optional[str]:
    """Собирает модель из токенов наименования [WORD] [NUMBER][LETTERS]."""
    skip = {"TSUYOKI", "ВОБЛЕР", "ЦВЕТ", "ЦВЕТА", "ЦВЕТУ"}
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if re.match(r"^[A-Za-z]+$", t) and t.upper() not in skip:
            if i + 1 < len(tokens):
                n = tokens[i + 1]
                if re.match(r"^\d+[A-Za-z]*$", n):
                    model = f"{t.upper()} {n.upper()}"
                    if i + 2 < len(tokens):
                        t2 = tokens[i + 2]
                        if _looks_like_color_code(t2):
                            return model
                        if re.match(r"^\d+[A-Za-z]*$", t2):
                            model = f"{t.upper()} {n.upper()} {t2.upper()}"
                            return model
                        if re.match(r"^[A-Za-z]+$", t2) and len(t2) <= 3:
                            model = f"{t.upper()} {n.upper()}{t2.upper()}"
                            return model
                    return model
            i += 1
        else:
            i += 1
    return None


_WEIGHT_SUFFIX_RE = re.compile(r"\s+\d+([,]\d+)?\s*гр\s*$", re.I)


def parse_name_to_model_and_code(name: str) -> tuple[Optional[str], Optional[str]]:
    """Из наименования товара извлекает модель и код цвета. Returns (model, color_code)."""
    if not name or not name.strip():
        return None, None
    s = name.strip()
    s = _WEIGHT_SUFFIX_RE.sub("", s).strip()
    for prefix in ("Воблер TsuYoki", "Воблер", "TsuYoki", "ВОБЛЕР TSUYOKI", "ВОБЛЕР"):
        if s.upper().startswith(prefix.upper()):
            s = s[len(prefix):].strip()
            break
    tokens = s.split()
    if not tokens:
        return None, None
    color_code = tokens[-1] if tokens else None
    model_tokens_raw = tokens[:-1] if len(tokens) > 1 else []
    skip_in_model = {"TSUYOKI", "ВОБЛЕР", "ЦВЕТ", "ЦВЕТА", "ЦВЕТУ"}
    model_tokens = [t for t in model_tokens_raw if t.upper() not in skip_in_model]
    model = " ".join(model_tokens).strip().upper() if model_tokens else None
    if model and color_code:
        return model, color_code.upper()
    if model:
        return model, None
    return None, color_code.upper() if color_code else None


def build_index(lures_root: Optional[Path] = None) -> dict:
    """Сканирует папку с моделями: base_model -> [(folder_path, [(filename, color_code), ...]), ...]"""
    root = lures_root or DEFAULT_LURES_ROOT
    if not root.is_dir():
        return {}
    t0 = time.perf_counter()
    logger.info("Сканирование локальной папки приманок: %s …", root)
    index = {}
    for folder in root.iterdir():
        if not folder.is_dir():
            continue
        bucket = _lure_model_index_key(folder.name)
        if not bucket:
            continue
        files_with_codes = []
        for f in folder.iterdir():
            if f.suffix.lower() not in (".jpg", ".jpeg", ".png", ".webp"):
                continue
            code = _extract_color_code_from_filename(f.name)
            files_with_codes.append((f.name, code))
        index.setdefault(bucket, []).append((str(folder), files_with_codes))
    n_models = len(index)
    n_files = sum(len(fc) for groups in index.values() for _, fc in groups)
    logger.info(
        "Локальный индекс готов: моделей %s, файлов изображений %s, за %.1f с",
        n_models,
        n_files,
        time.perf_counter() - t0,
    )
    return index


def find_local_image(
    model: str,
    color_code: str,
    lures_root: Optional[Path] = None,
    index: Optional[dict] = None,
) -> Optional[Path]:
    """Ищет файл в локальной папке по модели и коду цвета."""
    root = lures_root or DEFAULT_LURES_ROOT
    if index is None:
        index = build_index(root)
    model_norm = _normalize_model_for_folder(model)
    color_norm = (color_code or "").strip().upper()
    if not model_norm or not color_norm:
        return None

    folders = []
    for key in _lure_model_lookup_keys(model):
        folders = index.get(key, [])
        if folders:
            break
    for folder_path, files in folders:
        for filename, code in files:
            if code and (code == color_norm or color_norm in code or code in color_norm):
                return Path(folder_path) / filename
            if color_norm in (filename or "").upper():
                return Path(folder_path) / filename
    return None


def find_local_image_from_name(
    name: str,
    lures_root: Optional[Path] = None,
    index: Optional[dict] = None,
) -> Optional[Path]:
    """Из наименования извлекает модель и код цвета и ищет локальный файл."""
    model, code = parse_name_to_model_and_code(name)
    if not model:
        return None
    return find_local_image(model, code or "", lures_root=lures_root, index=index)


def build_remote_index(
    github_repo: str = DEFAULT_GITHUB_REPO,
    github_branch: str = DEFAULT_GITHUB_BRANCH,
    github_lures_dir: str = DEFAULT_GITHUB_LURES_DIR,
    session: Optional[requests.Session] = None,
    timeout: int = 30,
) -> dict:
    """Строит индекс изображений из GitHub-репозитория через API дерева."""
    index = {}
    repo = (github_repo or "").strip()
    branch = (github_branch or "").strip()
    lures_dir = (github_lures_dir or "").strip().strip("/")
    if not repo or not branch or not lures_dir:
        return index

    owner_repo = repo
    tree_url = f"https://api.github.com/repos/{owner_repo}/git/trees/{quote(branch, safe='')}?recursive=1"
    req_session = session or requests.Session()
    t0 = time.perf_counter()
    logger.debug("GitHub tree API: %s", tree_url)
    response = req_session.get(tree_url, timeout=timeout)
    response.raise_for_status()
    payload = response.json()
    tree_items = payload.get("tree", [])

    by_folder = {}
    prefix = f"{lures_dir}/"
    for item in tree_items:
        if item.get("type") != "blob":
            continue
        path = str(item.get("path") or "")
        if not path.startswith(prefix):
            continue
        lower = path.lower()
        if not lower.endswith((".jpg", ".jpeg", ".png", ".webp")):
            continue

        rel_path = path[len(prefix):]
        rel = PurePosixPath(rel_path)
        if len(rel.parts) < 2:
            continue
        folder_name = rel.parts[0]
        filename = rel.name
        color_code = _extract_color_code_from_filename(filename)
        encoded_path = quote(path, safe="/")
        raw_url = f"https://raw.githubusercontent.com/{owner_repo}/{quote(branch, safe='')}/{encoded_path}"
        by_folder.setdefault(folder_name, []).append((filename, color_code, raw_url))

    for folder_name, files in by_folder.items():
        bucket = _lure_model_index_key(folder_name)
        if not bucket:
            continue
        index.setdefault(bucket, []).append((folder_name, files))
    n_img = sum(len(files) for groups in index.values() for _, files in groups)
    logger.info(
        "Индекс GitHub: ключей моделей %s, файлов изображений %s, за %.1f с",
        len(index),
        n_img,
        time.perf_counter() - t0,
    )
    return index


def find_remote_image(
    model: str,
    color_code: str,
    index: Optional[dict] = None,
) -> Optional[tuple[str, str]]:
    """Ищет файл по модели/цвету в индексе GitHub. Returns (filename, raw_url)."""
    if index is None:
        return None
    model_norm = _normalize_model_for_folder(model)
    color_norm = (color_code or "").strip().upper()
    if not model_norm or not color_norm:
        return None

    folders = []
    for key in _lure_model_lookup_keys(model):
        folders = index.get(key, [])
        if folders:
            break
    for _folder_name, files in folders:
        for filename, code, raw_url in files:
            if code and (code == color_norm or color_norm in code or code in color_norm):
                return filename, raw_url
            if color_norm in (filename or "").upper():
                return filename, raw_url
    return None


def find_remote_image_from_name(
    name: str,
    index: Optional[dict] = None,
) -> Optional[tuple[str, str]]:
    """Из наименования извлекает модель/цвет и ищет в индексе GitHub."""
    model, code = parse_name_to_model_and_code(name)
    if not model:
        return None
    return find_remote_image(model, code or "", index=index)


# ---------------------------------------------------------------------------
# Выравнивание изображений воблеров (квадрат, по центру, горизонтально)
# ---------------------------------------------------------------------------

MIN_CONTOUR_AREA = 250
MIN_FOREGROUND_POINTS = 50  # минимум точек для PCA по маске (если контур не найден)
WHITE_THRESHOLDS = (245, 240, 235, 230, 220, 210, 200, 190, 180)  # светлые воблеры — низкие пороги
PADDING_RATIO = 0.12
MIN_REFINE_ANGLE_DEG = 0.2  # порог уточняющего поворота по fitLine (убирает остаточный наклон)


def _mask_foreground(img, white_threshold: int = 240):
    """Маска объекта: не-белый фон. Ожидается светлый/белый фон."""
    if img.ndim == 3:
        bg = (
            (img[:, :, 0] >= white_threshold)
            & (img[:, :, 1] >= white_threshold)
            & (img[:, :, 2] >= white_threshold)
        )
    else:
        bg = img >= white_threshold
    mask = np.where(bg, 0, 255).astype(np.uint8)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    return mask


def _mask_foreground_grayscale(img) -> Optional[np.ndarray]:
    """Запасная маска по яркости: объект темнее фона (для светлых воблеров на белом)."""
    if img.ndim == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        gray = img
    try:
        thresh_val, mask_bin = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    except Exception:
        mask_bin = (gray < 252).astype(np.uint8) * 255
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    mask = cv2.morphologyEx(mask_bin, cv2.MORPH_CLOSE, kernel)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    return mask


def _largest_contour(mask):
    """Контур с максимальной площадью."""
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return None
    return max(contours, key=cv2.contourArea)


def _cv_read_image(path: Path):
    """
    Надежное чтение изображений на Windows (в т.ч. пути с кириллицей/спецсимволами).
    1) cv2.imread
    2) np.fromfile + cv2.imdecode
    3) PIL fallback (RGB -> BGR)
    """
    path = Path(path)
    img = cv2.imread(str(path))
    if img is not None:
        return img
    try:
        data = np.fromfile(str(path), dtype=np.uint8)
        if data.size > 0:
            img = cv2.imdecode(data, cv2.IMREAD_COLOR)
            if img is not None:
                return img
    except Exception:
        pass
    if _PIL_AVAILABLE:
        try:
            pil_img = Image.open(path).convert("RGB")
            arr = np.array(pil_img, dtype=np.uint8)
            return cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
        except Exception:
            pass
    return None


def _cv_write_jpg(path: Path, img, quality: int = 95) -> bool:
    """
    Надежная запись JPG на Windows (в т.ч. пути с кириллицей/спецсимволами).
    """
    path = Path(path)
    ok = cv2.imwrite(str(path), img, [cv2.IMWRITE_JPEG_QUALITY, quality])
    if ok:
        return True
    try:
        ok, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, quality])
        if ok:
            buf.tofile(str(path))
            return True
    except Exception:
        pass
    return False


def make_square_centered(src_path: Path, dest_path: Path) -> bool:
    """Квадрат с белым фоном, картинка по центру (PIL). При неудаче align_lure_image."""
    if not _PIL_AVAILABLE:
        logger.debug("make_square_centered: PIL не установлен")
        return False
    try:
        src_path = Path(src_path)
        dest_path = Path(dest_path)
        if not src_path.is_file():
            return False
        img = Image.open(src_path).convert("RGB")
        w, h = img.size
        if w < 2 or h < 2:
            return False
        side = max(w, h)
        canvas = Image.new("RGB", (side, side), (255, 255, 255))
        ox = (side - w) // 2
        oy = (side - h) // 2
        canvas.paste(img, (ox, oy))
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        canvas.save(str(dest_path), "JPEG", quality=95)
        logger.info(f"Квадрат по центру (PIL): {src_path.name} -> {dest_path.name}")
        return True
    except Exception as e:
        logger.warning(f"make_square_centered {getattr(src_path, 'name', src_path)}: {e}")
        return False


def _pca_angle(points: np.ndarray) -> float:
    """Угол главной оси точек к оси X (радианы). Ось = направление макс. дисперсии."""
    if len(points) < 2:
        return 0.0
    pts = points.astype(np.float64)
    center = pts.mean(axis=0)
    centered = pts - center
    cov = np.cov(centered.T)
    eigenvalues, eigenvectors = np.linalg.eigh(cov)
    main_axis = eigenvectors[:, -1]
    angle = math.atan2(main_axis[1], main_axis[0])
    # Нормализуем к [-pi/2, pi/2]: длинная ось горизонтальна без переворота вверх ногами
    if angle > math.pi / 2:
        angle -= math.pi
    elif angle < -math.pi / 2:
        angle += math.pi
    return angle


def align_lure_image(
    src_path: Path,
    dest_path: Path,
    padding_ratio: float = PADDING_RATIO,
) -> bool:
    """Выравнивает воблер по горизонтали по оси PCA (контур или маска), квадрат с белым фоном."""
    try:
        if not _ALIGN_AVAILABLE:
            logger.debug("align_lure_image: opencv не установлен")
            return False
        src_path = Path(src_path)
        dest_path = Path(dest_path)
        if not src_path.is_file():
            logger.warning(f"align_lure_image: файл не найден {src_path}")
            return False
        img = _cv_read_image(src_path)
        if img is None:
            logger.warning(f"align_lure_image: не удалось прочитать {src_path}")
            return False
        h, w = img.shape[:2]
        if h < 10 or w < 10:
            return False

        points = None
        best_mask = None
        best_n = 0
        contour_points = None
        for thresh in WHITE_THRESHOLDS:
            mask = _mask_foreground(img, white_threshold=thresh)
            c = _largest_contour(mask)
            n_foreground = int(np.sum(mask != 0))
            if n_foreground > best_n:
                best_n = n_foreground
                best_mask = mask
            if c is not None and cv2.contourArea(c) >= MIN_CONTOUR_AREA:
                contour_points = c.reshape(-1, 2).astype(np.float64)
        # PCA по всем точкам маски (полный силуэт воблера) — ось совпадает с воблером, не с фрагментом контура
        if best_mask is not None and best_n >= MIN_FOREGROUND_POINTS:
            rows, cols = np.where(best_mask != 0)
            points = np.column_stack((cols.astype(np.float64), rows.astype(np.float64)))
        if points is None and contour_points is not None:
            points = contour_points

        if points is None or len(points) < MIN_FOREGROUND_POINTS:
            gmask = _mask_foreground_grayscale(img)
            if gmask is not None:
                n_g = int(np.sum(gmask > 0))
                if n_g >= MIN_FOREGROUND_POINTS:
                    rows, cols = np.where(gmask > 0)
                    points = np.column_stack((cols.astype(np.float64), rows.astype(np.float64)))

        if points is None or len(points) < MIN_FOREGROUND_POINTS:
            side = max(h, w)
            square = np.ones((side, side, 3), dtype=np.uint8) * 255
            ox = (side - w) // 2
            oy = (side - h) // 2
            square[oy : oy + h, ox : ox + w] = img
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            ok = _cv_write_jpg(dest_path, square, quality=95)
            if ok:
                logger.info(f"Квадрат без поворота (мало точек): {src_path.name} -> {dest_path.name}")
            return ok

        center_pt = points.mean(axis=0)
        angle_rad = _pca_angle(points)
        # В координатах изображения (Y направлена вниз) корректный поворот
        # для выравнивания оси по горизонтали имеет тот же знак, что и angle_rad.
        rotation_deg = math.degrees(angle_rad)

        M = cv2.getRotationMatrix2D((float(center_pt[0]), float(center_pt[1])), rotation_deg, 1.0)
        corners = np.array([[0, 0], [w, 0], [w, h], [0, h]], dtype=np.float32)
        rotated_corners = cv2.transform(corners.reshape(1, 4, 2), M).reshape(4, 2)
        rn_x, rn_y = rotated_corners[:, 0], rotated_corners[:, 1]
        nw = int(np.ceil(rn_x.max() - rn_x.min()))
        nh = int(np.ceil(rn_y.max() - rn_y.min()))
        M[0, 2] += -rn_x.min()
        M[1, 2] += -rn_y.min()
        img = cv2.warpAffine(img, M, (nw, nh), borderValue=(255, 255, 255))
        pts_rot = cv2.transform(points.reshape(1, -1, 2), M).reshape(-1, 2)
        h, w = img.shape[:2]

        x_min, y_min = pts_rot.min(axis=0)
        x_max, y_max = pts_rot.max(axis=0)
        bbox_w = x_max - x_min
        bbox_h = y_max - y_min
        # Длинная ось воблера должна быть горизонтальной (ширина bbox > высоты)
        if bbox_h > bbox_w:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            h, w = img.shape[:2]
            pts_rot = np.column_stack([h - 1 - pts_rot[:, 1], pts_rot[:, 0]])
            x_min, y_min = pts_rot.min(axis=0)
            x_max, y_max = pts_rot.max(axis=0)
            bbox_w, bbox_h = x_max - x_min, y_max - y_min

        # Уточнение угла по fitLine: ось строго горизонтальна
        pts_rot_32 = pts_rot.astype(np.float32)
        line = cv2.fitLine(pts_rot_32, cv2.DIST_L2, 0, 0.01, 0.01)
        line = np.asarray(line).ravel()
        vx, vy = float(line[0]), float(line[1])
        refine_rad = math.atan2(vy, vx)
        refine_deg = math.degrees(refine_rad)
        if abs(refine_deg) >= MIN_REFINE_ANGLE_DEG:
            cx, cy = (x_min + x_max) / 2, (y_min + y_max) / 2
            M_ref = cv2.getRotationMatrix2D((cx, cy), refine_deg, 1.0)
            img = cv2.warpAffine(img, M_ref, (w, h), borderValue=(255, 255, 255))
            cos_a = math.cos(math.radians(refine_deg))
            sin_a = math.sin(math.radians(refine_deg))
            dx = pts_rot[:, 0] - cx
            dy = pts_rot[:, 1] - cy
            pts_rot = np.column_stack([cx + dx * cos_a - dy * sin_a, cy + dx * sin_a + dy * cos_a])
            x_min, y_min = pts_rot.min(axis=0)
            x_max, y_max = pts_rot.max(axis=0)

        # Квадрат: всё повёрнутое изображение по центру, без обрезки по bbox (чтобы не резать лопатку/хвост)
        h, w = img.shape[:2]
        side = max(w, h)
        square = np.ones((side, side, 3), dtype=np.uint8) * 255
        ox = (side - w) // 2
        oy = (side - h) // 2
        square[oy : oy + h, ox : ox + w] = img
        dest_path.parent.mkdir(parents=True, exist_ok=True)
        ok = _cv_write_jpg(dest_path, square, quality=95)
        if ok:
            logger.info(f"Выравнивание применено: {src_path.name} -> {dest_path.name}")
        return ok
    except Exception as e:
        logger.warning(f"Ошибка выравнивания {src_path.name}: {e}")
        return False


# ---------------------------------------------------------------------------
# Парсер tsuyoki.ru
# ---------------------------------------------------------------------------

class TsuYokiParser:
    """Парсер для сайта tsuyoki.ru"""

    def __init__(
        self,
        input_file='tsuyokiarticles.xlsx',
        output_dir_ready='TsuYoki_ready',
        output_dir_site='TsuYoki_site',
        output_dir_raw='TsuYoki_raw',
        github_repo: str = DEFAULT_GITHUB_REPO,
        github_branch: str = DEFAULT_GITHUB_BRANCH,
        github_lures_dir: str = DEFAULT_GITHUB_LURES_DIR,
        use_local_lures: bool = False,
        lures_root: Optional[Path] = None,
    ):
        self.input_file = input_file
        self.output_dir_ready = Path(output_dir_ready)
        self.output_dir_site = Path(output_dir_site)
        self.output_dir_raw = Path(output_dir_raw)
        self.base_url = 'https://tsuyoki.ru'
        self.search_url_template = 'https://tsuyoki.ru/search/?q={article}&s=Поиск'
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        self.failed_items = []
        self.max_retries = 3
        self.use_local_lures = use_local_lures
        self.lures_root = (Path(lures_root) if lures_root is not None else DEFAULT_LURES_ROOT).resolve()
        self.github_repo = github_repo
        self.github_branch = github_branch
        self.github_lures_dir = github_lures_dir
        self._remote_index = None
        self._local_index = None
        self._create_output_dir()

    def _mark_failed(self, article: str, reason: str, source_file: Optional[str] = None):
        """Сохраняет информацию о неуспешной обработке для вывода в конце."""
        self.failed_items.append({
            "article": str(article).strip(),
            "reason": reason,
            "source_file": source_file or "-",
        })

    def _is_product_url(self, href):
        if not href:
            return False
        href = href.strip()
        if href == "/catalog/" or href.endswith("/catalog/"):
            return False
        if any(k in href.lower() for k in ['/product/', '/item/', '/goods/']):
            return True
        if '/catalog/' in href.lower():
            return href.count("/") >= 4
        return href.count("/") >= 4

    def _is_product_image_url(self, image_url):
        if not image_url:
            return False
        try:
            if image_url.startswith('/'):
                image_url = urljoin(self.base_url, image_url)
            parsed = urlparse(image_url)
            if not parsed.netloc.endswith("tsuyoki.ru"):
                return False
            if "/upload/" not in image_url:
                return False
            if not image_url.lower().endswith((".jpg", ".jpeg", ".png", ".webp")):
                return False
            return True
        except Exception:
            return False

    def _create_output_dir(self):
        for label, path in (
            ("готовые (выровненные)", self.output_dir_ready),
            ("с сайта", self.output_dir_site),
            ("исходники", self.output_dir_raw),
        ):
            try:
                path.mkdir(parents=True, exist_ok=True)
                logger.info(f"Папка {label}: {path}")
            except Exception as e:
                logger.error(f"Ошибка при создании папки {path}: {e}")
                raise

    def _save_raw_local(self, src_path: Path, basename: str) -> Optional[Path]:
        """Сохраняет исходный локальный файл в TsuYoki_raw."""
        try:
            suffix = src_path.suffix.lower() if src_path.suffix else ".jpg"
            raw_path = self.output_dir_raw / f"{basename}{suffix}"
            shutil.copy2(src_path, raw_path)
            return raw_path
        except Exception as e:
            logger.warning(f"Не удалось сохранить исходник (локальный): {e}")
            return None

    def _save_raw_downloaded(self, content: bytes, basename: str, image_url: str) -> Optional[Path]:
        """Сохраняет исходно скачанный файл в TsuYoki_raw."""
        try:
            suffix = Path(urlparse(image_url).path).suffix.lower() or ".jpg"
            if suffix not in (".jpg", ".jpeg", ".png", ".webp"):
                suffix = ".jpg"
            raw_path = self.output_dir_raw / f"{basename}{suffix}"
            with open(raw_path, "wb") as f:
                f.write(content)
            return raw_path
        except Exception as e:
            logger.warning(f"Не удалось сохранить исходник (скачанный): {e}")
            return None

    def _random_delay(self, min_seconds=2, max_seconds=4):
        delay = random.uniform(min_seconds, max_seconds)
        logger.debug(f"Пауза {delay:.2f} секунд")
        time.sleep(delay)

    def _make_request(self, url, retries=None):
        if retries is None:
            retries = self.max_retries
        for attempt in range(retries):
            try:
                response = self.session.get(url, timeout=30)
                response.raise_for_status()
                return response
            except requests.exceptions.RequestException as e:
                if attempt < retries - 1:
                    wait_time = (attempt + 1) * 2
                    logger.warning(f"Ошибка запроса (попытка {attempt + 1}/{retries}): {e}. Повтор через {wait_time} сек.")
                    time.sleep(wait_time)
                else:
                    logger.error(f"Не удалось выполнить запрос после {retries} попыток: {e}")
                    return None
        return None

    def _parse_search_results(self, html_content, article):
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            product_links = []
            search_containers = soup.find_all('div', class_=lambda x: x and 'search_page-iblock' in str(x))
            if search_containers:
                for container in search_containers:
                    links = container.find_all('a', href=True)
                    for link in links:
                        href = link.get('href', '')
                        if href and href != '/catalog/' and not href.endswith('/catalog/'):
                            product_links.append(link)
                    if product_links:
                        break
            if not product_links:
                selectors = [
                    'div.catalog-item a', 'div.product-item a', 'a.product-link', 'div.item a',
                    '.catalog a[href*="/product/"]', '.catalog a[href*="/item/"]',
                    'a[href*="/product/"]', 'a[href*="/item/"]', f'a[href*="{article}"]',
                ]
                for selector in selectors:
                    try:
                        links = soup.select(selector)
                        if links:
                            product_links = links
                            break
                    except Exception:
                        continue
            if not product_links:
                catalog_title = soup.find(string=lambda text: text and 'Каталог' in text and 'tsuyoki' in text.lower())
                if catalog_title:
                    catalog_container = catalog_title.find_parent('div', class_=lambda x: x and 'search_page-iblock' in str(x))
                    if catalog_container:
                        product_links = catalog_container.find_all('a', href=True)
                if not product_links:
                    all_links = soup.find_all('a', href=True)
                    product_links = [link for link in all_links
                                      if any(kw in link.get('href', '').lower() for kw in ['/product/', '/item/', '/goods/'])]
            for link in product_links:
                href = link.get('href', '')
                link_text = link.get_text(strip=True)
                if href and self._is_product_url(href):
                    return urljoin(self.base_url, href)
                if article in href and href.count("/") >= 4:
                    return urljoin(self.base_url, href)
            article_links = soup.find_all('a', href=True, string=lambda text: text and str(article) in str(text))
            if article_links:
                href = article_links[0].get('href', '')
                if href:
                    return urljoin(self.base_url, href)
            if product_links:
                href = product_links[0].get('href', '')
                if href and href.count("/") >= 3:
                    return urljoin(self.base_url, href)
            return None
        except Exception as e:
            logger.error(f"Ошибка при парсинге результатов поиска: {e}")
            return None

    def _parse_product_page(self, html_content, article):
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            page_article = None
            article_selectors = [
                'span.article', 'div.article', '[class*="article"]', '[id*="article"]',
            ]
            for selector in article_selectors:
                elements = soup.select(selector)
                for elem in elements:
                    text = elem.get_text(strip=True)
                    if 'артикул' in text.lower():
                        parts = text.split()
                        for i, part in enumerate(parts):
                            if 'артикул' in part.lower() and i + 1 < len(parts):
                                page_article = parts[i + 1].strip(':,')
                                break
                        if not page_article:
                            page_article = text.replace('Артикул', '').replace('артикул', '').strip(' :,')
                        break
                if page_article:
                    break
            if not page_article:
                tables = soup.find_all('table')
                for table in tables:
                    rows = table.find_all('tr')
                    for row in rows:
                        cells = row.find_all(['td', 'th'])
                        if len(cells) >= 2 and 'артикул' in cells[0].get_text().lower():
                            page_article = cells[1].get_text(strip=True)
                            break
                    if page_article:
                        break
            if not page_article:
                dl_lists = soup.find_all('dl')
                for dl in dl_lists:
                    dts, dds = dl.find_all('dt'), dl.find_all('dd')
                    for dt, dd in zip(dts, dds):
                        if 'артикул' in dt.get_text().lower():
                            page_article = dd.get_text(strip=True)
                            break
                    if page_article:
                        break
            article_matched = False
            if page_article:
                page_article_clean = page_article.strip().lower().strip('{}')
                article_clean = str(article).strip().lower().strip('{}')
                if page_article_clean == article_clean:
                    article_matched = True
                else:
                    page_text = soup.get_text()
                    if article_clean in page_text.lower():
                        article_matched = True
            else:
                page_text = soup.get_text()
                article_clean = str(article).strip().lower().strip('{}')
                if article_clean in page_text.lower():
                    article_matched = True
            if not article_matched:
                return None, None

            image_url = None
            priority_images = soup.find_all('img', class_=lambda x: x and 'image_for_watch' in x and 'js_picture_glass' in x)
            if priority_images:
                img = priority_images[0]
                src = img.get('src') or img.get('data-src') or img.get('data-lazy-src')
                if src:
                    full_image_url = urljoin(self.base_url, src)
                    if self._is_product_image_url(full_image_url):
                        image_url = full_image_url
            if not image_url:
                image_selectors = [
                    ('img.image_for_watch.js_picture_glass',), ('img.js_picture_glass',), ('img.image_for_watch',),
                    ('img.product-image',), ('img.main-image',), ('div.product-image img',), ('div.main-image img',),
                    ('img[class*="product"]',), ('img[class*="main"]',), ('div.image img',),
                    ('.gallery img:first-child',), ('.product-gallery img:first-child',),
                ]
                for (selector,) in image_selectors:
                    try:
                        images = soup.select(selector)
                        if images:
                            img = images[0]
                            src = img.get('src') or img.get('data-src') or img.get('data-lazy-src')
                            if src:
                                potential_url = urljoin(self.base_url, src)
                                if self._is_product_image_url(potential_url):
                                    image_url = potential_url
                                    break
                    except Exception:
                        continue
            if not image_url:
                all_images = soup.find_all('img')
                for img in all_images:
                    src = img.get('src') or img.get('data-src') or img.get('data-lazy-src')
                    if src and not any(skip in src.lower() for skip in ['icon', 'logo', 'banner', 'ad']):
                        potential_url = urljoin(self.base_url, src)
                        if self._is_product_image_url(potential_url):
                            image_url = potential_url
                            break
            if not image_url:
                return None, None
            return page_article, image_url
        except Exception as e:
            logger.error(f"Ошибка при парсинге страницы товара: {e}")
            return None, None

    def _download_image(self, image_url, article, output_basename=None):
        try:
            response = self._make_request(image_url)
            if not response:
                return False
            basename = output_basename if output_basename is not None else article
            self._save_raw_downloaded(response.content, basename, image_url)
            file_path = self.output_dir_site / f"{basename}.jpg"
            with open(file_path, 'wb') as f:
                f.write(response.content)
            logger.info(f"Изображение сохранено: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Ошибка при скачивании изображения {image_url}: {e}")
            return False

    def _process_article(self, article, name=None, output_basename=None):
        article = str(article).strip()
        if not article:
            return False
        basename = output_basename if output_basename is not None else article
        logger.info(f"Обработка артикула: {article}" + (f" ({name[:50]}...)" if name and len(name) > 50 else (f" ({name})" if name else "")) + (f" -> {basename}.jpg" if output_basename else ""))

        if name and name.strip():
            nm = name.strip()
            if self.use_local_lures:
                if self._local_index is None:
                    self._local_index = build_index(self.lures_root)
                local_path = find_local_image_from_name(nm, lures_root=self.lures_root, index=self._local_index)
                if local_path and local_path.is_file():
                    dest = self.output_dir_ready / f"{basename}.jpg"
                    try:
                        self._save_raw_local(local_path, basename)
                        if align_lure_image(local_path, dest):
                            logger.info(f"Изображение из локальной папки выровнено и сохранено: {local_path.name} -> {dest}")
                            return True
                        diagnose_msg = "Локальный файл найден, но выравнивание не удалось"
                        if _cv_read_image(local_path) is None:
                            diagnose_msg = (
                                "Локальный файл не читается (cv2/PIL). "
                                "Проверьте путь/кодировку имени файла и целостность файла"
                            )
                        logger.warning(f"{diagnose_msg}: {local_path}")
                        self._mark_failed(article, diagnose_msg, str(local_path))
                        return False
                    except Exception as e:
                        logger.warning(f"Не удалось обработать локальный файл: {e}")
                        self._mark_failed(article, f"Ошибка обработки локального файла: {e}", local_path.name)
                        return False
                logger.debug("В локальной папке не найдено изображение по наименованию, пробуем GitHub")

            if self._remote_index is None:
                try:
                    self._remote_index = build_remote_index(
                        github_repo=self.github_repo,
                        github_branch=self.github_branch,
                        github_lures_dir=self.github_lures_dir,
                        session=self.session,
                    )
                    logger.info(f"Индекс изображений GitHub загружен: {self.github_repo}@{self.github_branch}")
                except Exception as e:
                    logger.warning(f"Не удалось собрать индекс изображений из GitHub: {e}")
                    self._remote_index = {}

            remote_match = find_remote_image_from_name(nm, index=self._remote_index)
            if remote_match:
                filename, remote_url = remote_match
                dest = self.output_dir_ready / f"{basename}.jpg"
                temp_path: Optional[Path] = None
                try:
                    response = self._make_request(remote_url)
                    if not response:
                        self._mark_failed(article, "Не удалось скачать файл изображения из GitHub", remote_url)
                        return False
                    self._save_raw_downloaded(response.content, basename, remote_url)
                    suffix = Path(urlparse(remote_url).path).suffix.lower() or ".jpg"
                    if suffix not in (".jpg", ".jpeg", ".png", ".webp"):
                        suffix = ".jpg"
                    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                        tmp.write(response.content)
                        temp_path = Path(tmp.name)
                    if align_lure_image(temp_path, dest):
                        logger.info(f"Изображение из GitHub выровнено и сохранено: {filename} -> {dest}")
                        return True

                    diagnose_msg = "Файл из GitHub найден, но выравнивание не удалось"
                    if temp_path is None or _cv_read_image(temp_path) is None:
                        diagnose_msg = "Файл из GitHub не читается (cv2/PIL). Проверьте целостность файла"
                    logger.warning(f"{diagnose_msg}: {remote_url}")
                    self._mark_failed(article, diagnose_msg, remote_url)
                    return False
                except Exception as e:
                    logger.warning(f"Не удалось обработать файл из GitHub: {e}")
                    self._mark_failed(article, f"Ошибка обработки файла из GitHub: {e}", remote_url)
                    return False
                finally:
                    if temp_path and temp_path.exists():
                        try:
                            temp_path.unlink()
                        except Exception:
                            pass
            else:
                logger.debug("В GitHub-репозитории не найдено изображение по наименованию, ищем на сайте")

        search_url = self.search_url_template.format(article=f"{{{article}}}")
        logger.info(f"Поиск по URL: {search_url}")
        response = self._make_request(search_url)
        if not response:
            logger.error(f"Не удалось выполнить поиск для артикула {article}")
            self._mark_failed(article, "Не удалось выполнить поиск на сайте")
            return False
        self._random_delay()
        product_url = self._parse_search_results(response.text, article)
        if not product_url:
            logger.warning(f"Товар не найден для артикула {article} на странице поиска")
            self._mark_failed(article, "Товар не найден на странице поиска")
            return False
        self._random_delay()
        logger.info(f"Переход на страницу товара: {product_url}")
        response = self._make_request(product_url)
        if not response:
            logger.error(f"Не удалось загрузить страницу товара для артикула {article}")
            self._mark_failed(article, "Не удалось загрузить страницу товара")
            return False
        self._random_delay()
        page_article, image_url = self._parse_product_page(response.text, article)
        if not image_url:
            logger.warning(f"Не найдено изображение для артикула {article}")
            self._mark_failed(article, "На странице товара не найдено изображение")
            return False
        if not self._download_image(image_url, article, output_basename=basename):
            self._mark_failed(article, "Не удалось скачать изображение")
            return False
        self._random_delay()
        logger.info(f"Успешно обработан артикул: {article}")
        return True

    def read_articles_from_excel(self, filename=None, sheet_name=0, column=0):
        if filename is None:
            filename = self.input_file
        articles = []
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filename, read_only=True, data_only=True)
            if isinstance(sheet_name, int):
                sheet = workbook.worksheets[sheet_name]
            else:
                sheet = workbook[sheet_name]
            col_index = column + 1
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
                cell_value = row[0].value
                if cell_value is not None:
                    article = str(cell_value).strip()
                    if article and article.lower() != 'nan':
                        articles.append(article)
            workbook.close()
            logger.info(f"Прочитано {len(articles)} артикулов из {filename}")
            return articles
        except FileNotFoundError:
            logger.error(f"Файл {filename} не найден")
            return []
        except ImportError as e:
            logger.error(f"Установите openpyxl: pip install openpyxl. Ошибка: {e}")
            return []
        except Exception as e:
            logger.error(f"Ошибка при чтении файла {filename}: {e}")
            return []

    def read_offers_from_xml(self, xml_source):
        pairs = []
        gr_pattern = re.compile(r"GR-\d+")
        try:
            if xml_source.strip().lower().startswith(("http://", "https://")):
                response = self._make_request(xml_source)
                if not response:
                    logger.error(f"Не удалось загрузить XML по URL: {xml_source}")
                    return []
                response.encoding = response.apparent_encoding or "utf-8"
                root = ET.fromstring(response.text)
            else:
                tree = ET.parse(xml_source)
                root = tree.getroot()
            for offer in root.findall(".//offer"):
                url_el = offer.find("url")
                url_text = (url_el.text or "").strip() if url_el is not None else ""
                gr_match = gr_pattern.search(url_text)
                if not gr_match:
                    continue
                gr_code = gr_match.group(0)
                name_el = offer.find("name")
                article = None
                for param in offer.findall("param"):
                    if param.get("name") == "Артикул" and param.text:
                        article = param.text.strip()
                        break
                name = (name_el.text or "").strip() if name_el is not None else ""
                if article and "tsuyoki" in name.lower():
                    pairs.append((gr_code, article, name))
            logger.info(f"Прочитано {len(pairs)} офферов TsuYoki из XML")
        except ET.ParseError as e:
            logger.error(f"Ошибка разбора XML: {e}")
        except FileNotFoundError:
            logger.error(f"Файл не найден: {xml_source}")
        return pairs

    def process_from_xml(self, xml_path):
        pairs = self.read_offers_from_xml(xml_path)
        if not pairs:
            logger.error("Список офферов из XML пуст")
            return
        logger.info(f"Начало обработки {len(pairs)} офферов из XML")
        for i, (gr_code, article, name) in enumerate(pairs, 1):
            logger.info(f"Прогресс: {i}/{len(pairs)}")
            self._process_article(article, name=name, output_basename=gr_code)
        self._log_completion(len(pairs))

    def _log_completion(self, total):
        logger.info(f"\nОбработка завершена! Всего: {total}, пропущено: {len(self.failed_items)}")
        if self.failed_items:
            logger.info("Необработанные позиции:")
            for item in self.failed_items:
                logger.info(
                    f"  - Артикул: {item['article']} | Причина: {item['reason']} | Файл: {item['source_file']}"
                )

    def process_all(self, sheet_name=0, column=0):
        articles = self.read_articles_from_excel(sheet_name=sheet_name, column=column)
        if not articles:
            logger.error("Список артикулов пуст")
            return
        logger.info(f"Начало обработки {len(articles)} артикулов")
        for i, article in enumerate(articles, 1):
            logger.info(f"Прогресс: {i}/{len(articles)}")
            self._process_article(article)
        self._log_completion(len(articles))


def main():
    import argparse
    if os.name == "nt":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    base_output_dir = Path("TsuYoki_images")
    parser = argparse.ArgumentParser(description='Парсер для сайта tsuyoki.ru')
    parser.add_argument('-i', '--input', default='tsuyokiarticles.xlsx',
                        help='Путь к Excel файлу со списком артикулов (по умолчанию: tsuyokiarticles.xlsx)')
    parser.add_argument('--output-ready', default=str(base_output_dir / 'TsuYoki_ready'),
                        help='Папка для готовых изображений (по умолчанию: TsuYoki_images/TsuYoki_ready)')
    parser.add_argument('-o', '--output-site', dest='output_site', default=str(base_output_dir / 'TsuYoki_site'),
                        help='Папка для изображений, скачанных с сайта (по умолчанию: TsuYoki_images/TsuYoki_site)')
    parser.add_argument('--output-raw', dest='output_raw', default=str(base_output_dir / 'TsuYoki_raw'),
                        help='Папка для исходников без выравнивания (по умолчанию: TsuYoki_images/TsuYoki_raw)')
    parser.add_argument('-s', '--sheet', type=int, default=0,
                        help='Индекс листа в Excel файле (по умолчанию: 0)')
    parser.add_argument('-c', '--column', type=int, default=0,
                        help='Индекс столбца с артикулами (по умолчанию: 0)')
    parser.add_argument('--xml', nargs='?', const='https://gria.ru/bitrix/catalog_export/imageless_offers.xml',
                        metavar='URL_или_файл',
                        help='Источник офферов: URL или путь к XML. По умолчанию — XML с gria.ru.')
    parser.add_argument('--excel', action='store_true',
                        help='Брать артикулы из Excel (-i/--input) вместо XML.')
    parser.add_argument('--github-repo', default=DEFAULT_GITHUB_REPO,
                        help='Репозиторий GitHub с папкой изображений (по умолчанию: netebla/TsuYoki_Parser)')
    parser.add_argument('--github-branch', default=DEFAULT_GITHUB_BRANCH,
                        help='Ветка GitHub для поиска изображений (по умолчанию: main)')
    parser.add_argument('--github-lures-dir', default=DEFAULT_GITHUB_LURES_DIR,
                        help='Путь к каталогу моделей в репозитории (по умолчанию: TsuYoki Lures 2014-2026)')
    parser.add_argument('--local-lures', action='store_true',
                        help='Сначала искать изображения в локальной папке (см. --lures-root), затем GitHub, затем сайт')
    parser.add_argument('--lures-root', default=None, metavar='ПУТЬ',
                        help='Корень каталога с подпапками моделей для --local-lures (по умолчанию: TsuYoki Lures 2014-2026 рядом с main.py)')
    parser.add_argument('-v', '--verbose', action='store_true',
                        help='Подробный лог (DEBUG) в консоль')
    parser.add_argument('--log-file', default=None, metavar='ФАЙЛ',
                        help='Дублировать лог в файл UTF-8 (создаёт каталоги при необходимости)')
    args = parser.parse_args()
    configure_runtime_logging(
        log_file=Path(args.log_file).expanduser() if args.log_file else None,
        verbose=args.verbose,
    )
    parser_instance = TsuYokiParser(
        input_file=args.input,
        output_dir_ready=args.output_ready,
        output_dir_site=args.output_site,
        output_dir_raw=args.output_raw,
        github_repo=args.github_repo,
        github_branch=args.github_branch,
        github_lures_dir=args.github_lures_dir,
        use_local_lures=args.local_lures,
        lures_root=Path(args.lures_root) if args.lures_root else None,
    )
    xml_source = args.xml if args.xml is not None else 'https://gria.ru/bitrix/catalog_export/imageless_offers.xml'
    if args.excel:
        parser_instance.process_all(sheet_name=args.sheet, column=args.column)
    else:
        parser_instance.process_from_xml(xml_source)


if __name__ == '__main__':
    main()


